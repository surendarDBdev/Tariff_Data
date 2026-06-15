import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def create_pgvcl_workbook():
    wb = openpyxl.Workbook()
    ws_glossary = wb.active
    ws_glossary.title = "Glossary & Metadata"
    ws_glossary.views.sheetView[0].showGridLines = True

    font_family = "Segoe UI"
    title_font = Font(name=font_family, size=16, bold=True, color="D35400")
    subtitle_font = Font(name=font_family, size=11, italic=True, color="595959")
    section_font = Font(name=font_family, size=13, bold=True, color="D35400")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=10, bold=False, color="000000")
    bold_data_font = Font(name=font_family, size=10, bold=True, color="000000")
    
    header_fill = PatternFill(start_color="D35400", end_color="D35400", fill_type="solid")
    zebra_fill = PatternFill(start_color="FFF5EE", end_color="FFF5EE", fill_type="solid")
    sub_category_fill = PatternFill(start_color="FFE4D1", end_color="FFE4D1", fill_type="solid")
    
    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    wrap_left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin_border_side = Side(border_style="thin", color="E0D0C0")
    medium_border_side = Side(border_style="medium", color="D35400")
    
    cell_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    header_border = Border(left=thin_border_side, right=thin_border_side, top=medium_border_side, bottom=medium_border_side)

    def adjust_column_widths(ws):
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 10)

    # Title Block
    ws_glossary.cell(row=2, column=2, value="PGVCL Gujarat Electricity Tariff Model").font = title_font
    ws_glossary.cell(row=3, column=2, value="Paschim Gujarat Vij Company Limited (GERC Approved)").font = subtitle_font
    
    # 1. Abbreviations Section
    ws_glossary.cell(row=5, column=2, value="1. Acronyms & Abbreviations Reference").font = section_font
    headers_glossary = ["Abbreviation", "Expanded Term", "Domain/Context", "Definition / Operational Applicability"]
    for col_idx, header in enumerate(headers_glossary, start=2):
        cell = ws_glossary.cell(row=6, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = left_align

    glossary_data = [
        ("GERC", "Gujarat Electricity Regulatory Commission", "Regulatory Commission", "The state-level statutory authority setting retail power tariffs in Gujarat."),
        ("PGVCL", "Paschim Gujarat Vij Company Limited", "Distribution Utility", "State-owned distribution licensee serving Saurashtra and Kutch regions of Gujarat."),
        ("RGP", "Residential General Purpose", "Tariff Class", "Low-tension residential billing schedule."),
        ("Non-RGP", "Non-Residential General Purpose", "Tariff Class", "Low-tension commercial/non-domestic billing schedule."),
        ("HTMD", "High Tension Maximum Demand", "Voltage Supply", "High tension supply above 100 kVA (kVAh basis)."),
        ("FPPAS", "Fuel & Power Purchase Adjustment Surcharge", "Variable Surcharge", "Pass-through quarterly adjustment based on actual generation fuel costs.")
    ]
    
    curr_row = 7
    for row_data in glossary_data:
        for col_idx, val in enumerate(row_data, start=2):
            cell = ws_glossary.cell(row=curr_row, column=col_idx, value=val)
            cell.font = data_font
            cell.border = cell_border
            if col_idx == 2:
                cell.font = bold_data_font
            cell.alignment = left_align
            if curr_row % 2 == 1:
                cell.fill = zebra_fill
        curr_row += 1

    # 2. Metadata Section (Effective Date, DISCOMs)
    curr_row += 2
    ws_glossary.cell(row=curr_row, column=2, value="2. General Tariff Framework & Metadata").font = section_font
    curr_row += 1

    general_notes = [
        ("Effective Date", "1st April 2025 (Approved via GERC Retail Supply Tariff Order)"),
        ("Included DISCOMs", "Paschim Gujarat Vij Company Limited (PGVCL)"),
        ("Billing Frequency", "Monthly/Bi-Monthly based on consumer categories."),
        ("Rounding Rules", "Billing parameters are rounded off to the nearest complete Rupee.")
    ]

    ws_glossary.cell(row=curr_row, column=2, value="Regulatory Parameter").font = header_font
    ws_glossary.cell(row=curr_row, column=2).fill = header_fill
    ws_glossary.cell(row=curr_row, column=2).border = header_border
    ws_glossary.cell(row=curr_row, column=3, value="Operational Rule Description").font = header_font
    ws_glossary.cell(row=curr_row, column=3).fill = header_fill
    ws_glossary.cell(row=curr_row, column=3).border = header_border
    ws_glossary.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=5)

    curr_row += 1
    for param, desc in general_notes:
        cell_p = ws_glossary.cell(row=curr_row, column=2, value=param)
        cell_p.font = bold_data_font
        cell_p.border = cell_border
        cell_p.alignment = left_align
        
        ws_glossary.merge_cells(start_row=curr_row, start_column=3, end_row=curr_row, end_column=5)
        for col_idx in range(3, 6):
            cell_d = ws_glossary.cell(row=curr_row, column=col_idx)
            cell_d.border = cell_border
            if col_idx == 3:
                cell_d.value = desc
                cell_d.font = data_font
                cell_d.alignment = wrap_left_align
        curr_row += 1

    adjust_column_widths(ws_glossary)

    # -------------------------------------------------------------
    # TAB 2: TARIFF SCHEDULE
    # -------------------------------------------------------------
    ws_tariff = wb.create_sheet(title="Tariff Schedule")
    ws_tariff.views.sheetView[0].showGridLines = True
    
    ws_tariff.append([])
    ws_tariff.cell(row=2, column=2, value="Base Slabs, Subsidies, and Net Rates").font = title_font
    ws_tariff.cell(row=3, column=2, value="GERC Approved Base Slabs for FY 2025-26").font = subtitle_font
    ws_tariff.append([])
    ws_tariff.append([])
    
    ws_tariff.cell(row=6, column=2, value="Energy Tariffs and Fixed Demand Charges").font = section_font
    
    headers_tariff = [
        "Major Category", 
        "Sub-Category / Slab Description", 
        "Applicability Range", 
        "Fixed / Demand Charges (₹/kW or HP/month)", 
        "Base Energy Charge (₹/unit)",
        "Government Subsidy (₹/unit)",
        "Wheeling Surcharge (₹/unit)",
        "Net Effective Energy Charge (₹/unit)"
    ]
        
    for col_idx, header in enumerate(headers_tariff, start=2):
        cell = ws_tariff.cell(row=7, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = left_align if col_idx <= 4 else right_align
        cell.border = header_border
        
    tariff_data = [
        # [Category, Sub-Category, Applicability, Fixed_Charge, Base_Energy, Subsidy, Wheeling]
        ("Residential (RGP - LT)", "Urban Slab 0 - 50 units", "Urban domestic connections", 15.00, 3.05, 0.00, 0.00),
        ("Residential (RGP - LT)", "Urban Slab 51 - 100 units", "Urban domestic connections", 25.00, 3.50, 0.00, 0.00),
        ("Residential (RGP - LT)", "Urban Slab 101 - 250 units", "Urban domestic connections", 45.00, 4.15, 0.00, 0.00),
        ("Residential (RGP - LT)", "Urban Slab Above 250 units", "Urban domestic connections", 70.00, 5.15, 0.00, 0.00),
        ("Residential (RGP - LT-Rural)", "Rural Slab 0 - 50 units", "Rural domestic connections", 15.00, 2.65, 0.00, 0.00),
        ("Residential (RGP - LT-Rural)", "Rural Slab 51 - 100 units", "Rural domestic connections", 25.00, 3.10, 0.00, 0.00),
        ("Residential (RGP - LT-Rural)", "Rural Slab 101 - 250 units", "Rural domestic connections", 45.00, 3.75, 0.00, 0.00),
        ("Residential (RGP - LT-Rural)", "Rural Slab Above 250 units", "Rural domestic connections", 70.00, 4.90, 0.00, 0.00),
        ("Commercial (Non-RGP LT)", "Commercial LT Supply", "Shops and offices under 50 kW", 100.00, 4.50, 0.00, 0.00),
        ("HT Industrial", "HTMD-1 (11 kV Supply)", "HT industries with load > 100 kVA", 170.00, 4.00, 0.00, 0.00),
        ("HT Industrial", "HTMD-1 (EHT Supply)", "HT industries with load > 100 kVA", 285.00, 3.85, 0.00, 0.00),
        ("Agriculture (LT)", "Irrigation Pump Sets", "Farming irrigation under LT", 20.00, 0.60, 0.00, 0.00)
    ]
        
    curr_row = 8
    prev_category = ""
    for row_data in tariff_data:
        category = row_data[0]
        ws_tariff.cell(row=curr_row, column=2, value=category).font = bold_data_font if category != prev_category else data_font
        ws_tariff.cell(row=curr_row, column=3, value=row_data[1]).font = data_font
        ws_tariff.cell(row=curr_row, column=4, value=row_data[2]).font = data_font
        
        # Fixed Charges
        c_fixed = ws_tariff.cell(row=curr_row, column=5, value=row_data[3])
        c_fixed.font = data_font
        c_fixed.number_format = '"₹"#,##0.00'
        c_fixed.alignment = right_align
        
        # Base Energy Charge
        c_base = ws_tariff.cell(row=curr_row, column=6, value=row_data[4])
        c_base.font = data_font
        c_base.number_format = '"₹"#,##0.00'
        c_base.alignment = right_align
        
        # Government Subsidy
        c_sub = ws_tariff.cell(row=curr_row, column=7, value=row_data[5])
        c_sub.font = data_font
        c_sub.number_format = '"₹"#,##0.00'
        c_sub.alignment = right_align
        
        # Wheeling Charge
        c_wheel = ws_tariff.cell(row=curr_row, column=8, value=row_data[6])
        c_wheel.font = data_font
        c_wheel.number_format = '"₹"#,##0.00'
        c_wheel.alignment = right_align
        
        # Formula: Net = Base - Subsidy + Wheeling
        c_net = ws_tariff.cell(row=curr_row, column=9, value=f"=F{curr_row}-G{curr_row}+H{curr_row}")
        c_net.font = bold_data_font
        c_net.number_format = '"₹"#,##0.00'
        c_net.alignment = right_align
            
        for c in range(2, 10):
            ws_tariff.cell(row=curr_row, column=c).border = cell_border
            if category != prev_category:
                ws_tariff.cell(row=curr_row, column=c).fill = sub_category_fill
                
        prev_category = category
        curr_row += 1
        
    adjust_column_widths(ws_tariff)

    # -------------------------------------------------------------
    # TAB 3: SURCHARGES & RULES
    # -------------------------------------------------------------
    ws_surcharges = wb.create_sheet(title="Surcharges & Rules")
    ws_surcharges.views.sheetView[0].showGridLines = True
    
    ws_surcharges.append([])
    ws_surcharges.cell(row=2, column=2, value="Electricity Surcharges, Penalties, and Adjustments").font = title_font
    ws_surcharges.cell(row=3, column=2, value="Additional Financial Liabilities, Power Factor, and Special Rules").font = subtitle_font
    ws_surcharges.append([])
    ws_surcharges.append([])
    
    ws_surcharges.cell(row=6, column=2, value="Statutory Taxes, Surcharges & Operational Penalties").font = section_font
    
    surcharge_headers = ["Charge/Adjustment Type", "Rate / Quantum", "Base of Application", "Applicability & Regulatory Rules"]
    for col_idx, header in enumerate(surcharge_headers, start=2):
        cell = ws_surcharges.cell(row=7, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = left_align
    
    rules_data = [
        ("Electricity Duty (ED)", "15.00%", "Total active energy charges", "State tax collected for Gujarat Government."),
        ("Delayed Payment Surcharge", "1.50% per month", "Outstanding arrears past due date", "Billed on simple interest daily basis."),
        ("Overdrawal Penalty", "Double standard Fixed Charge", "Actual demand exceeding contract limit", "Surcharge levied on excess load drawn."),
        ("Power Factor Surcharge", "2.00% on Energy Charge", "Average power factor dropping below 0.85 lag", "Applicable on commercial and industrial categories.")
    ]
    
    curr_row = 8
    for row_data in rules_data:
        ws_surcharges.cell(row=curr_row, column=2, value=row_data[0]).font = bold_data_font
        ws_surcharges.cell(row=curr_row, column=3, value=row_data[1]).font = data_font
        ws_surcharges.cell(row=curr_row, column=4, value=row_data[2]).font = data_font
        ws_surcharges.cell(row=curr_row, column=5, value=row_data[3]).font = data_font
        
        for c in range(2, 6):
            ws_surcharges.cell(row=curr_row, column=c).border = cell_border
        
        if curr_row % 2 == 1:
            for c in range(2, 6):
                ws_surcharges.cell(row=curr_row, column=c).fill = zebra_fill
        curr_row += 1
        
    adjust_column_widths(ws_surcharges)

    wb.save("PGVCL_Gujarat_Electricity_Tariff_Order_Model.xlsx")
    print("[SUCCESS] Compiled: PGVCL_Gujarat_Electricity_Tariff_Order_Model.xlsx")

if __name__ == "__main__":
    create_pgvcl_workbook()
